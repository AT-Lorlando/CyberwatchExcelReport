from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
from charts import ChartGenerator
from utils import get_legend_df

MATURITY_LEVELS = {
    "high": 3,
    "functional": 2,
    "proof-of-concept": 1,
    "unproven": 0,
}


class ReportGenerator:
    # A constant dictionary to map colums to their types, like numeric or string, verbose or not, etc.
    CRITICITY_COLORS = {
        "C1": "FFFF7575",  # Red
        "C2": "FFFABF8F",  # Orange
        "C3": "FFFFE575",  # Yellow
        "C4": "FF00B050",  # Green
        "C5": "FF5A8AC6",  # Blue (default)
    }
    PRIORITY_COLORS = {
        "P1": "FFF8696B",  # Red
        "P2": "FFfa9395",  # Orange
        "P3": "FFfcc6c7",  # Yellow
        "P4": "FFcad9ed",  # Green
        "P5": "FF8fafd8",  # Blue (default)
        "P6": "FF5A8AC6",  # Blue (default)
    }
    MATURITY_COLORS = {
        "high": "FFFF7575",  # Red
        "functional": "FFFEC4B8",  # Orange
        "proof-of-concept": "FFFFE575",  # Yellow
        "unproven": "FFFFF6C9",  # Green
    }
    COLOR_SCALE = {
        "min": "FF5A8AC6",
        "max": "FFF8696B",
        "mid": "FFFFFFFF",
    }
    STYLE = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )

    def __init__(
        self,
        data_df,
        cpe_df,
        patch_df,
        issue_df,
        old_cve_dfs,
        score_col="CVSS Computed Score",
        groupby=["CVE Code", "Server"],
        date_format="%Y-%m-%d",
    ):
        self.dataframe = data_df
        self.cpe_df = cpe_df
        self.patch_df = patch_df
        self.issue_df = issue_df
        self.old_cve_dfs = old_cve_dfs
        self.score_col = score_col
        self.groupby = groupby
        self.date_format = date_format
        self.get_sheets()

    def get_sheets(self):
        for df in [self.dataframe, self.cpe_df, self.patch_df, self.issue_df]:
            parse_dataframe(df, format=self.date_format)
        self.cpe_df.sort_values(by="Total", ascending=False, inplace=True)
        self.patch_df.sort_values(by="CVE Number", ascending=False, inplace=True)

        for i, df in enumerate(self.old_cve_dfs):
            parse_dataframe(df, format=self.date_format)

        for i, df in enumerate(self.old_cve_dfs):
            if (
                "Status" not in self.old_cve_dfs[i].columns
            ):  # If the status column is present, the dataframe is already computed
                compute_dataframe(
                    self.old_cve_dfs[i],
                    self.old_cve_dfs[i + 1] if i + 1 < len(self.old_cve_dfs) else None,
                    score_col=self.score_col,
                    groupby=self.groupby,
                )

        compute_dataframe(
            self.dataframe,
            self.old_cve_dfs[0] if len(self.old_cve_dfs) > 0 else None,
            score_col=self.score_col,
            groupby=self.groupby,
        )
        self.cve_df = group_df(self.dataframe, self.score_col, groupby=self.groupby)
        self.sheets = {
            "CVE Scan": self.cve_df,
            "CPE Scan": self.cpe_df,
            "Patch Scan": self.patch_df,
            "Security issues Scan": self.issue_df,
            "Data": self.dataframe,
        }
        for i, old_cve_df in enumerate(self.old_cve_dfs, start=1):
            self.sheets.update({f"old n{i} CVE Scan": old_cve_df})

    def apply_conditional_formatting(
        self, ws, df, color_scale_columns=["CVSS Computed Score"]
    ):
        if not isinstance(color_scale_columns, list):
            color_scale_columns = [color_scale_columns]

        def apply_cell_rule(column_name, criteria_colors):
            if column_name in df.columns:
                col_letter = get_column_letter(df.columns.get_loc(column_name) + 1)
                for criterion, color in criteria_colors.items():
                    ws.conditional_formatting.add(
                        f"{col_letter}2:{col_letter}{ws.max_row}",
                        CellIsRule(
                            operator="equal",
                            formula=[f'"{criterion}"'],
                            stopIfTrue=True,
                            fill=PatternFill(
                                start_color=color, end_color=color, fill_type="solid"
                            ),
                        ),
                    )

        def apply_color_scale(column_name, start_value, mid_value, end_value):
            if column_name in df.columns:
                col_letter = get_column_letter(df.columns.get_loc(column_name) + 1)
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws.max_row}",
                    ColorScaleRule(
                        start_type="percent",
                        start_value=start_value,
                        start_color=self.COLOR_SCALE["min"],
                        mid_type="num",
                        mid_value=mid_value,
                        mid_color=self.COLOR_SCALE["mid"],
                        end_type="percent",
                        end_value=end_value,
                        end_color=self.COLOR_SCALE["max"],
                    ),
                )

        def apply_number_format(column_name, number_format):
            if column_name in df.columns:
                col_letter = get_column_letter(df.columns.get_loc(column_name) + 1)
                for row in range(2, ws.max_row + 1):
                    ws[f"{col_letter}{row}"].number_format = number_format

        for col_num in range(1, df.shape[1] + 1):
            ws[f"{get_column_letter(col_num)}1"].font = Font(color="FFFFFF")

        apply_cell_rule("Criticity", self.CRITICITY_COLORS)
        apply_cell_rule("Priority", self.PRIORITY_COLORS)
        apply_cell_rule("Maturity", self.MATURITY_COLORS)

        for col_name in color_scale_columns:
            apply_color_scale(col_name, 1, 4, 100)

        apply_color_scale("Score EPSS", 1, 0.05, 100)
        apply_color_scale("Total", 1, 20, 100)
        apply_color_scale("CVE Number", 1, 20, 100)
        apply_cell_rule("Cisa Reference", {"Yes": "00B050"})
        apply_number_format("Update CVSS", "+0.00;-0.00;0.00")
        apply_number_format("Update EPSS", "+0.00%;-0.00%;0.00%")

        for date_col in ["Published Date", "Last Reviewed Date"]:
            apply_number_format(date_col, "YYYY/MM/DD")

        if "Update Cisa" in df.columns:
            col_letter = get_column_letter(df.columns.get_loc("Update Cisa") + 1)
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{ws.max_row}",
                CellIsRule(
                    operator="equal",
                    formula=['"Added"'],
                    stopIfTrue=True,
                    fill=PatternFill(
                        start_color="00B050", end_color="00B050", fill_type="solid"
                    ),
                ),
            )
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{ws.max_row}",
                CellIsRule(
                    operator="equal",
                    formula=['"Removed"'],
                    stopIfTrue=True,
                    fill=PatternFill(
                        start_color="FF7575", end_color="FF7575", fill_type="solid"
                    ),
                ),
            )

        for update_col in ["Update EPSS", "Update CVSS"]:
            if update_col in df.columns:
                col_letter = get_column_letter(df.columns.get_loc(update_col) + 1)
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws.max_row}",
                    CellIsRule(
                        operator="greaterThan",
                        formula=[0],
                        stopIfTrue=True,
                        font=Font(color="FF7575"),
                    ),
                )
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws.max_row}",
                    CellIsRule(
                        operator="lessThan",
                        formula=[0],
                        stopIfTrue=True,
                        font=Font(color="00B050"),
                    ),
                )

        max_width = 25
        for col in df.columns:
            col_letter = get_column_letter(df.columns.get_loc(col) + 1)
            max_length = max(
                len(str(ws[f"{col_letter}{row}"].value))
                for row in range(1, min(50, ws.max_row + 1))
            )
            ws.column_dimensions[col_letter].width = (
                min(max_length, max_width) + 5
                if not col.startswith("Description")
                else max_length + 5
            )

    def apply_charts(self, ws, by_scans=True):
        chart_generator = ChartGenerator(self.dataframe, self.old_cve_dfs, "charts")
        images = [
            chart_generator.generate_cwe_chart(),
            chart_generator.generate_capec_chart(),
            # chart_generator.generate_cve_by_group_chart(
            #     group_columns=["Server", "Priority"]
            # ),
            chart_generator.generate_cve_by_group_chart(
                group_columns=["Domain", "Server", "Priority"]
            ),
            chart_generator.generate_mean_cvss_by_group_chart(
                group_column="Domain", score_col=self.score_col
            ),
            chart_generator.generate_mean_cvss_by_group_chart(
                group_column="Server", score_col=self.score_col
            ),
            chart_generator.generate_criticity_by_group_chart(group_column="Domain"),
            chart_generator.generate_criticity_by_group_chart(group_column="Server"),
            chart_generator.generate_priority_by_group_chart(group_column="Domain"),
            chart_generator.generate_priority_by_group_chart(group_column="Server"),
            chart_generator.generate_cve_by_date_chart(),
            chart_generator.generate_cve_by_scan_chart() if by_scans else None,
            (
                chart_generator.generate_mean_cvss_by_scan_chart(self.score_col)
                if by_scans
                else None
            ),
        ]
        idx = 0
        for image in images:
            if not image:
                continue
            img = Image(image)
            col = 1 + (idx // 2) * 11
            row = 1 + (idx % 2) * 25
            letter = get_column_letter(col)
            ws.add_image(img, f"{letter}{row}")
            idx += 1

    def add_table_from_df(self, ws: Worksheet, df: pd.DataFrame, name):
        table = Table(
            displayName=name.replace(" ", "_"),
            ref=f"A1:{get_column_letter(df.shape[1])}{len(df)+1}",
        )
        table.tableStyleInfo = self.STYLE
        ws.add_table(table)

    def write_to_excel_with_loader(self, df, sheet_name, writer):
        from tqdm import tqdm

        chunk_size = 10000
        with tqdm(total=len(df), desc="Writing to Excel", unit="rows") as pbar:
            for start_row in range(0, len(df), chunk_size):
                end_row = min(start_row + chunk_size, len(df))
                df.iloc[start_row:end_row].to_excel(
                    writer,
                    sheet_name=sheet_name,
                    startrow=start_row,
                    index=False,
                    header=start_row == 0,
                )
                pbar.update(chunk_size)

    def write_to_excel(self, df, sheet_name, writer):
        if df is not None:
            print(f"Writing {sheet_name} sheet")
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    def generate_report(self, filename, date):
        filename = f"{filename}_{date}.xlsx"
        self.sheets = {"Legende": get_legend_df(date), **self.sheets}
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            for sheet_name, df in self.sheets.items():
                self.write_to_excel(df, sheet_name, writer)
            print("Saving file (this may take a while)")

        try:
            wb = load_workbook(filename)
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return

        for sheet_name, df in self.sheets.items():
            print(f"Applying formatting to {sheet_name} sheet")
            try:
                ws = wb[sheet_name]
                print(ws)
                self.add_table_from_df(ws, df, sheet_name)
            except:
                print(f"Sheet {sheet_name} not found")
            if df is not None:
                ws = wb[sheet_name]
                self.apply_conditional_formatting(
                    ws, df, color_scale_columns=[self.score_col]
                )

        graph_sheet = wb.create_sheet("Analysis")
        self.apply_charts(graph_sheet, by_scans=len(self.sheets) > 5)
        print("Saving file (this may take a while)")
        wb.save(filename)

    def generate_synthesis(
        self,
        filename,
        date,
        subset=["Server", "CVE Code", "Product"],
        groupby=["CVE Code", "Server", "Status"],
    ):
        """
        Generate a single excel file with all the CVE from every scan in a single sheet
        if a cve is present in multiple scans, the latest scan will be kept
        """
        filename = f"{filename}_{date}.xlsx"
        synthesis_df = pd.concat([self.dataframe] + self.old_cve_dfs)
        groupby = [col for col in groupby if col in list(synthesis_df.columns)]
        # If a CVE is in the old scan but not in the self.dataframe, it means it's been fixed
        # The comparison is done on the rows CVE Code, Server, and Product
        # Set its status to "Fixed"
        keys = ["CVE Code", "Server", "Product"]
        for i, old_cve_df in enumerate(self.old_cve_dfs):
            fixed_cves = old_cve_df[
                ~old_cve_df.set_index(keys).index.isin(
                    self.dataframe.set_index(keys).index
                )
            ]
            fixed_cves_index = synthesis_df.set_index(keys).index.isin(
                fixed_cves.set_index(keys).index
            )
            synthesis_df.loc[fixed_cves_index, "Status"] = "Fixed"

        synthesis_df = synthesis_df.drop_duplicates(subset=subset, keep="first")
        synthesis_df = group_df(synthesis_df, self.score_col, groupby=groupby)

        with pd.ExcelWriter(
            filename, engine="openpyxl", date_format="dd/mm/yyyy"
        ) as writer:
            synthesis_df.to_excel(writer, sheet_name="Synthesis", index=False)
        wb = load_workbook(filename)
        ws = wb["Synthesis"]
        self.apply_conditional_formatting(
            ws, synthesis_df, color_scale_columns=[self.score_col]
        )
        self.add_table_from_df(ws, synthesis_df, "Synthesis")
        wb.save(filename)


def get_news_from_scans_vectorized(
    merged_df: pd.DataFrame, score_col: str
) -> pd.DataFrame:
    merged_df["Status"] = "Known"
    update_cols = ["Update Cisa", "Update EPSS", "Update CVSS", "Update Maturity"]
    merged_df[update_cols] = ""

    # Define conditions for updates
    update_conditions = [
        merged_df["Cisa Reference"] != merged_df["Cisa Reference_old"],
        merged_df["Maturity"] != merged_df["Maturity_old"],
        merged_df["Score EPSS"] != merged_df["Score EPSS_old"],
        merged_df[score_col] != merged_df[score_col + "_old"],
    ]

    # Apply the conditions to set the status
    merged_df.loc[pd.concat(update_conditions, axis=1).any(axis=1), "Status"] = (
        "Updated"
    )
    merged_df.loc[merged_df[score_col + "_old"].isna(), "Status"] = "New"
    updated_df = merged_df.loc[~merged_df[score_col + "_old"].isna()]

    if "Cisa Reference" in updated_df.columns:
        cisa_condition = (
            updated_df["Cisa Reference"] != updated_df["Cisa Reference_old"]
        )
        updated_df.loc[cisa_condition, "Update Cisa"] = updated_df.loc[
            cisa_condition, "Cisa Reference"
        ].map({"Yes": "Added", "No": "Removed"})

    if "Score EPSS" in updated_df.columns:
        score_diff = updated_df["Score EPSS"] - updated_df["Score EPSS_old"]
        updated_df.loc[score_diff != 0, "Update EPSS"] = score_diff

    if score_col in updated_df.columns:
        custom_score_diff = updated_df[score_col] - updated_df[score_col + "_old"]
        updated_df.loc[custom_score_diff != 0, "Update CVSS"] = custom_score_diff

    if "Maturity" in updated_df.columns:
        maturity_diff = updated_df["Maturity"].map(MATURITY_LEVELS) - updated_df[
            "Maturity_old"
        ].map(MATURITY_LEVELS)
        updated_df.loc[maturity_diff != 0, "Update Maturity"] = (
            updated_df["Maturity_old"] + " -> " + updated_df["Maturity"]
        )

    merged_df[update_cols] = updated_df[update_cols]
    return merged_df


def compute_dataframe(
    dataframe, old_df, score_col, groupby=["CVE Code", "Server"]
) -> pd.DataFrame:
    # Add a new column to the dataframe to calculate the priority
    dataframe["Priority"] = 6
    conditions = [
        dataframe["Cisa Reference"] == "Yes",
        dataframe["Maturity"] == "high",
        dataframe["Score EPSS"] >= 0.8,
        dataframe[score_col] >= 9,
    ]
    dataframe.loc[dataframe[score_col] >= 7, "Priority"] -= 1
    for condition in conditions:
        dataframe.loc[condition, "Priority"] = 5
    for condition in conditions:
        dataframe.loc[condition, "Priority"] -= 1

    dataframe["Priority"] = "P" + dataframe["Priority"].astype(str)

    # Compare old and actual scan
    if old_df is not None:
        merged_df = dataframe.merge(
            old_df,
            on=["Server", "CVE Code", "Component"],
            suffixes=("", "_old"),
            how="left",
        )

        merged_df = get_news_from_scans_vectorized(merged_df, score_col)

        dataframe["Status"] = merged_df["Status"]
        update_cols = [col for col in merged_df.columns if col.startswith("Update")]
        print(dataframe)
        dataframe.loc[:, update_cols] = merged_df.loc[:, update_cols]


def group_df(dataframe, score_col, groupby=["CVE Code", "Server"]):
    clone = dataframe.copy()
    base_agg = {
        # Join the servers names with a pipe, but once by servers names (no duplicates)
        "Server": lambda x: " | ".join(x.astype(str).unique()),
        "Component": lambda x: " | ".join(x.astype(str).unique()),
        "Product": lambda x: " | ".join(x.astype(str).unique()),
        "Version": lambda x: " | ".join(x.astype(str).unique()),
        "Patch": lambda x: " | ".join(x.astype(str)),
        "Criticity": "first",
        "Priority": "first",
        "Score EPSS": "first",
        score_col: "first",
        "Maturity": "first",
        "Cisa Reference": "first",
        "Priority": "first",
        "Status": "first",
        "Update Cisa": "first",
        "Update EPSS": "first",
        "Update CVSS": "first",
        "Update Maturity": "first",
    }
    score_agg = {
        "CVSS Score": "first",
        "CVSS Environmental Score": "first",
        "CVSS Temporal Score": "first",
        "CVSS Computed Score": "first",
    }
    score_agg.pop(score_col)
    optional_agg = {
        "Surface": "first",
        "Vector": "first",
        "Environmental Vector": "first",
        "Temporal Vector": "first",
        "Related CWEs": "first",
        "Related CAPECs": "first",
        "Related ATK": "first",
        "CertFR References": "first",
    }
    useless_cols = [
        "Criticity",
        "Surface",
        "Vector",
        "Environmental Vector",
        "Temporal Vector",
        "Related CWEs",
        "Related CAPECs",
        "Related ATK",
        "CertFR References",
        "CVSS Score",
        "CVSS Temporal Score",
        "CVSS Computed Score",
        "CVSS Environmental Score",
        "Product",
        "Version",
    ]
    useless_cols.remove(score_col)
    agg = {}
    for obj in [base_agg, score_agg, optional_agg]:
        for col, rule in obj.items():
            if (
                col in list(clone.columns)
                and col not in useless_cols
                and col not in groupby
            ):
                agg[col] = rule
    clone = clone.groupby(groupby).agg(agg).reset_index()
    clone.sort_values(
        by=["Priority", "Score EPSS"], ascending=[True, False], inplace=True
    )
    return clone


def parse_dataframe(df, format="%Y-%m-%d"):
    if df.empty:
        df.loc[0, df.columns[0]] = "No data to display"
        return
    numeric_cols = [
        "CVSS Score",
        "CVSS Temporal Score",
        "CVSS Environmental Score",
        "CVSS Computed Score",
        "Score EPSS",
    ]
    replace_dict = {"Undefined": 0, "None": 0, pd.NA: 0}
    numeric_cols = [col for col in numeric_cols if col in list(df.columns)]
    for col in numeric_cols:
        df[col] = df[col].replace(replace_dict)
        df[col].fillna(0)

    replace_dict = {pd.NA: "None"}
    if "Patch" in list(df.columns):
        df["Patch"] = df["Patch"].replace(replace_dict)

    df[numeric_cols] = df[numeric_cols].astype(float)
    date_cols = [
        "Published Date",
        "Last Reviewed Date",
    ]
    date_cols = [col for col in date_cols if col in list(df.columns)]
    for col in date_cols:
        df[col] = pd.to_datetime(df[col], errors="coerce", format=format)
        df[col] = df[col].fillna("Unknown")
